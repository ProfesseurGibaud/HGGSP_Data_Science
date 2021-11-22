# -*- coding: utf-8 -*-
"""
Created on Thu Aug  5 14:29:11 2021

@author: Sylgi
"""

import openpyxl

import sqlite3
conn = sqlite3.connect('RéponseQuestionnaires.db')
cursor = conn.cursor()
# 111 lignes au total dans les réponses

modif1 = 0
modif2 = 0

response_workbook = openpyxl.load_workbook(r"reponses.xlsx")
response_sheet = response_workbook.active
#print(response_sheet.cell(row = 1,column = 1).value)

indication_workbook = openpyxl.load_workbook(r"indications.xlsx")
indication_sheet = indication_workbook.active
#print(indication_sheet.cell(row = 1,column = 1).value)

def debut_fin(number):
    # On ajoute + 1 pour avoir le bon numéro de ligne et de colonne
    if type(number) != str:
        number = str(number).split(".")
        debut = int(number[0]) + modif1
        fin = int(number[1]) + modif1
    else:
        liste = number.split(",")
        debut = int(liste[0]) + modif1
        fin = int(liste[1]) + modif1
    return debut,fin


def dans_set(test,doublé):
    debut = doublé[0]
    fin = doublé[1]
    return test in range(debut,fin+1)

def donner_bonne_question(valeur,dico_numeroligne_question):
    for doublé in dico_numeroligne_question:
        #print(doublé)
        if dans_set(valeur,doublé):
            return doublé,dico_numeroligne_question[doublé]


dico_numeroligne_question = {}
for row in range(1,13):
    #print(debut_fin(indication_sheet.cell(row = row,column = 1).value))
    #print(indication_sheet.cell(row = row,column = 2).value)
    dico_numeroligne_question[(debut_fin(indication_sheet.cell(row = row,column = 1).value))] = indication_sheet.cell(row = row,column = 2).value

dico_response = {}
for row in range(1,13):
    liste = (indication_sheet.cell(row = row,column = 3).value).split(",")
    for string in liste:
        num_string = string.split(":")
        #print(num_string)
        if int(num_string[0]) !=41:
            dico_response[int(num_string[0])] =num_string[1]



def make_dico_user(dico_response,response_sheet,dico_numeroligne_question,row_number):
    dico_user = {}
    for colonne in range(1,112):
        #print(response_sheet.cell(row = 1,column = colonne).value)
        if colonne not in dico_response.keys():
            dico_user[response_sheet.cell(row = 1,column = colonne).value] = response_sheet.cell(row = row_number,column = colonne).value
        else:
            doublé,bonne_question = donner_bonne_question(colonne, dico_numeroligne_question)
            if colonne == doublé[0]:
                bonne_reponse = "N/A"
                dico_user[bonne_question] = ""

            if colonne <= doublé[1]:
                #print(response_sheet.cell(row = 2,column = colonne).value)
                #print(response_sheet.cell(row = 2,column = colonne).value == "Oui")
                if response_sheet.cell(row = row_number,column = colonne).value == "Oui":
                    bonne_reponse = dico_response[colonne]
                    dico_user[bonne_question] += "/" + bonne_reponse
                    #print(bonne_question,dico_user[bonne_question])
                #print("  ")
    return dico_user

def insert_dans_db(dico_user):
    cursor.execute("""INSERT INTO Reponse('ID de la réponse',
     'Date de soumission',
     'Dernière page',
     'Langue de départ',
     'Tête de série',
     'Date de lancement',
     'Date de la dernière action',
     'Vous vous identifiez comme :',
     'Renseignez votre âge.',
     "Sélectionnez l'académie dans laquelle vous exercez.",
     'Renseignez votre statut actuel',
     'Renseignez votre statut actuel [Autre]',
     "Si vous avez obtenu un concours, précisez son année d'obtention",
     'parcours_académique',
     'Vous enseignez en :',
     'Renseignez le nom de votre lycée et la ville dans laquelle il se situe. Si vous exercez dans plusieurs établissements, merci de ne renseigner que celui dans lequel vous effectuez la majorité de votre service.',
     'Votre établissement propose la spécialité histoire-géographie, géopolitique, sciences politiques :',
     'enseignement spé :',
     'formation enseignement spé',
     'formation reçue',
     'se considère comme formé',
     'A ce jour, vous considérez vous comme formé.e à la géopolitique et à son enseignement ? [Non, pas assez]',
     'Vous enseignez la spécialité HGGSP en :',
     "thèmes ou vous êtes le plus à l'aise",
     "thèmes où vous êtes le moins à l'aise",
     'méthodologies',
     'Décrivez les côtés positifs de cet enseignement.',
     'Décrivez les côtés négatifs de cet enseignement.',
     'L’enseignement de la spécialité influe-t-il sur votre enseignement en tronc commun ?',
     'Si oui, précisez comment.',
     'Comment définissez-vous la géopolitique ?',
     'Etablissez-vous une distinction claire entre géopolitique et relations internationales ?',
     'Si oui, précisez comment cela se traduit dans la conception et dans la mise en œuvre de votre enseignement.',
     'Renseignez le pourcentage d’élèves ayant choisi la spécialité HGGSP en classe de Première dans votre établissement :',
     'motivation choix spé élèves première',
     'thèmes favoris des élèves en première',
     'Renseignez le pourcentage d’élèves ayant conservé la spécialité HGGSP en classe de Terminale dans votre établissement.',
     'motivation conservation spé en terminale',
     'thèmes favoris des élèves en terminale',
     'Souhaitez-vous être tenu informé.e des résultats de cette enquête ?',
     "Souhaitez-vous participer à un entretien individuel sur cette thématique ?  Vos impressions, vécus et retours sur l'enseignement de la géopolitique me sont essentiels pour avancer plus aisément dans mes recherches. N'hésitez pas à participer !",
     'Pour que je puisse vous tenir informé.e des résultats de mes recherches et/ou pour que je puisse vous contatcer pour organiser un entretien, merci de renseigner votre adresse mail.') VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", tuple(dico_user.values()))




def Fin():
    conn.commit()
    conn.close()


for row_number in range(3,819):
    dico_user = make_dico_user(dico_response, response_sheet, dico_numeroligne_question, row_number)
    insert_dans_db(dico_user)

